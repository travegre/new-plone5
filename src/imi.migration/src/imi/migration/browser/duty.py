# -*- coding: utf-8 -*-
from datetime import date
from datetime import datetime
from datetime import timedelta
from io import BytesIO

from plone import api
from Products.Five import BrowserView
from Products.Five.browser.pagetemplatefile import ViewPageTemplateFile


ROSTER_TYPE = 'imi.duty.roster_day'
ROSTER_FOLDER_ID = 'dezurstva-1'
COPY_FIELDS = (
    'dezurni_zdravnik', 'nadzorni_zdravnik', 'specializant', 'sarsifon',
    'dezurni_tehnik', 'izobrazevanje', 'bakterioloski_tehnik',
    'inoqula_tehnik', 'sprejem_predpriprava_tehnik', 'vpis',
    'intervencijska_skupina', 'intervencijska_skupina_telefon',
    'BOR', 'HIV', 'HUM', 'IT', 'PRZ', 'KLM', 'KOV', 'VINVZV', 'VINDIF',
    'WHO', 'kiestra', 'klukca',
)

TEAM_FIELDS = (
    ('dezurni_zdravnik', u'dežurni zdravnik'),
    ('nadzorni_zdravnik', u'nadzorni zdravnik'),
    ('specializant', u'specializant na usposabljanju'),
    ('sarsifon', u'Virološki laboratoriji'),
    ('bakterioloski_tehnik', u'Urgentni laboratorij'),
    ('inoqula_tehnik', u'BMK laboratorij'),
    ('sprejem_predpriprava_tehnik', u'sprejem/predpriprava'),
    ('vpis', u'vpis'),
)

READINESS_FIELDS = (
    ('BOR', u'BOR'),
    ('HIV', u'HIV'),
    ('HUM', u'HUM'),
    ('PRZ', u'PRZ'),
    ('IT', u'IT'),
    ('KLM', u'KLM'),
    ('KOV', u'VZ'),
    ('WHO', u'WHO'),
    ('kiestra', u'Kiestra'),
)

DAY_NAMES = {
    0: u'ponedeljek', 1: u'torek', 2: u'sreda', 3: u'četrtek',
    4: u'petek', 5: u'sobota', 6: u'nedelja',
}
MONTH_NAMES = {
    1: u'januar', 2: u'februar', 3: u'marec', 4: u'april',
    5: u'maj', 6: u'junij', 7: u'julij', 8: u'avgust',
    9: u'september', 10: u'oktober', 11: u'november', 12: u'december',
}


def _portal(context):
    return api.portal.get()


def _roster_folder(context):
    portal = _portal(context)
    folder = portal.get(ROSTER_FOLDER_ID)
    if folder is None:
        raise KeyError('Missing roster folder /%s/%s' % (portal.getId(), ROSTER_FOLDER_ID))
    return folder


def _normalise_date(value):
    value = (value or '').strip()
    if not value:
        raise ValueError('Datum je obvezen.')
    try:
        parsed = datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        raise ValueError('Datum mora biti veljaven in v obliki LLLL-MM-DD.')
    return parsed.strftime('%Y-%m-%d')


def _publish_if_possible(obj):
    try:
        transitions = api.content.get_transitions(obj=obj)
        ids = [item.get('id') for item in transitions]
        if 'publish' in ids:
            api.content.transition(obj=obj, transition='publish')
    except Exception:
        pass


class DutyPublicView(BrowserView):
    """Public Dežurstva frontend preserving the Plone-4 skin semantics."""

    template = ViewPageTemplateFile('duty_public.pt')

    def __call__(self):
        return self.template()

    @property
    def portal(self):
        return _portal(self.context)

    @property
    def roster_folder(self):
        return _roster_folder(self.context)

    def selected_date(self):
        raw = str(self.request.form.get('datum') or '').strip()
        if raw:
            for fmt in ('%Y-%m-%d', '%d.%m.%Y'):
                try:
                    return datetime.strptime(raw, fmt).date()
                except ValueError:
                    pass
        return date.today()

    def selected_date_id(self):
        return self.selected_date().strftime('%Y-%m-%d')

    def selected_date_sl(self):
        return self.selected_date().strftime('%d.%m.%Y')

    def formatted_date(self):
        selected = self.selected_date()
        prefix = u'danes, ' if selected == date.today() else u''
        return u'%s%s, %02d.%02d.%04d' % (
            prefix, DAY_NAMES[selected.weekday()], selected.day,
            selected.month, selected.year)

    def previous_date(self):
        return (self.selected_date() - timedelta(days=1)).strftime('%Y-%m-%d')

    def next_date(self):
        return (self.selected_date() + timedelta(days=1)).strftime('%Y-%m-%d')

    def roster(self):
        return self.roster_folder.get(self.selected_date_id())

    def _staff_directory(self):
        return self.portal.get('seznam_zaposlenih')

    def staff_title(self, employee_id):
        directory = self._staff_directory()
        employee_id = str(employee_id or '')
        if directory is not None:
            employee = directory.get(employee_id)
            if employee is not None:
                try:
                    return employee.Title() or employee_id
                except Exception:
                    pass
        return employee_id

    def _field_rows(self, definitions):
        roster = self.roster()
        if roster is None:
            return []
        rows = []
        for field_name, label in definitions:
            values = list(getattr(roster, field_name, ()) or ())
            if field_name == 'vpis':
                values.reverse()
            for employee_id in values:
                rows.append({
                    'field': field_name,
                    'label': label,
                    'employee_id': str(employee_id),
                    'name': self.staff_title(employee_id),
                })
        return rows

    def team_rows(self):
        return self._field_rows(TEAM_FIELDS)

    def readiness_rows(self):
        return self._field_rows(READINESS_FIELDS)

    def intervention_rows(self):
        roster = self.roster()
        if roster is None:
            return []
        return [self.staff_title(value)
                for value in (getattr(roster, 'intervencijska_skupina', ()) or ())]

    def intervention_phone(self):
        roster = self.roster()
        if roster is None:
            return ''
        return getattr(roster, 'intervencijska_skupina_telefon', '') or '040 190 945'

    def last_change(self):
        try:
            brains = self.portal.portal_catalog(
                portal_type=ROSTER_TYPE,
                path='/'.join(self.roster_folder.getPhysicalPath()),
                sort_on='modified', sort_order='descending')
            if not brains:
                return u''
            value = brains[0].modified.asdatetime()
            return value.strftime('%d.%m.%Y ob %H:%M')
        except Exception:
            return u''

    def staff_options(self):
        directory = self._staff_directory()
        if directory is None:
            return []
        result = []
        for employee in directory.objectValues():
            if getattr(employee, 'portal_type', None) != 'imi.staff.employee':
                continue
            if bool(getattr(employee, 'exclude_from_nav', False)):
                continue
            employee_id = str(employee.getId())
            result.append((employee_id, employee.Title() or employee_id))
        return sorted(result, key=lambda item: item[1].lower())

    def selected_person(self):
        return str(self.request.form.get('oseba') or '')

    def selected_person_title(self):
        return self.staff_title(self.selected_person()) if self.selected_person() else ''

    def person_dates(self):
        person = self.selected_person()
        if not person:
            return []
        rows = []
        try:
            brains = self.portal.portal_catalog(
                portal_type=ROSTER_TYPE,
                path='/'.join(self.roster_folder.getPhysicalPath()),
                sort_on='id', sort_order='ascending')
            for brain in brains:
                obj = brain.getObject()
                if person in tuple(getattr(obj, 'dezurni_zdravnik', ()) or ()):
                    try:
                        parsed = datetime.strptime(obj.getId(), '%Y-%m-%d').date()
                    except ValueError:
                        continue
                    rows.append({
                        'id': obj.getId(),
                        'date': parsed,
                        'year': parsed.year,
                        'month': parsed.month,
                        'label': '%02d.%02d.%04d' % (parsed.day, parsed.month, parsed.year),
                    })
        except Exception:
            pass
        return rows

    def person_dates_by_year(self):
        grouped = {}
        for item in self.person_dates():
            grouped.setdefault(item['year'], {}).setdefault(item['month'], []).append(item)
        result = []
        for year in sorted(grouped, reverse=True):
            months = []
            for month in sorted(grouped[year], reverse=True):
                months.append({
                    'number': month,
                    'name': MONTH_NAMES[month],
                    'dates': sorted(grouped[year][month], key=lambda x: x['date']),
                })
            result.append({'year': year, 'months': months})
        return result

    def easyform_url(self):
        form = self.portal.get('spremeni-dezurstvo')
        return form.absolute_url() if form is not None else ''


class DutyExportView(BrowserView):
    """Excel export replacement for the legacy ``izvoz.py`` flow."""

    def _parse_period(self):
        request = self.request.form
        mode = str(request.get('izbira') or request.get('izvoz-1-izbira') or request.get('izvoz-2-izbira') or '')
        today = date.today()
        monday = today - timedelta(days=today.weekday())
        if mode == 'teden':
            return monday - timedelta(days=7), monday - timedelta(days=1)
        if mode == 'tedenplus2':
            return monday, monday + timedelta(days=14)
        if mode == 'tekoci':
            first = today.replace(day=1)
            next_month = (first.replace(day=28) + timedelta(days=4)).replace(day=1)
            return first, next_month - timedelta(days=1)
        if mode == 'pretekli':
            first_this = today.replace(day=1)
            last_previous = first_this - timedelta(days=1)
            return last_previous.replace(day=1), last_previous
        if mode == 'oddo':
            start = str(request.get('od') or '').strip()
            end = str(request.get('do') or '').strip()
            return datetime.strptime(start, '%d.%m.%Y').date(), datetime.strptime(end, '%d.%m.%Y').date()
        raise ValueError('Izberite obdobje za izvoz.')

    def __call__(self):
        from openpyxl import Workbook

        first_day, last_day = self._parse_period()
        person = str(self.request.form.get('oseba') or self.request.form.get('izvoz-2-oseba') or '')
        portal = _portal(self.context)
        folder = _roster_folder(self.context)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'izpis'
        heading = 'Izpis: %s - %s' % (
            first_day.strftime('%d.%m.%Y'), last_day.strftime('%d.%m.%Y'))
        if person:
            directory = portal.get('seznam_zaposlenih')
            employee = directory.get(person) if directory is not None else None
            heading += ' za osebo %s' % (employee.Title() if employee is not None else person)
        sheet.cell(row=1, column=1, value=heading)

        row = 4
        brains = portal.portal_catalog(
            portal_type=ROSTER_TYPE,
            path='/'.join(folder.getPhysicalPath()),
            sort_on='id', sort_order='ascending')
        for brain in brains:
            try:
                day = datetime.strptime(brain.id, '%Y-%m-%d').date()
            except ValueError:
                continue
            if day < first_day or day > last_day:
                continue
            obj = brain.getObject()
            values = tuple(getattr(obj, 'dezurni_zdravnik', ()) or ())
            for employee_id in values:
                if person and employee_id != person:
                    continue
                directory = portal.get('seznam_zaposlenih')
                employee = directory.get(employee_id) if directory is not None else None
                name = employee.Title() if employee is not None else employee_id
                row += 1
                sheet.cell(row=row, column=1, value=day.strftime('%d.%m.%Y'))
                sheet.cell(row=row, column=2, value=u'dežurni zdravnik')
                sheet.cell(row=row, column=3, value=name)

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        filename = 'izpis_%s-%s.xlsx' % (
            first_day.strftime('%d_%m_%Y'), last_day.strftime('%d_%m_%Y'))
        response = self.request.response
        response.setHeader('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response.setHeader('Content-Disposition', 'attachment; filename="%s"' % filename)
        return payload.getvalue()


class DutyAdminView(BrowserView):
    """Plone 5 replacement for the legacy ``dezurstva_admin.pt`` page."""


class DutyHomeView(BrowserView):
    """Default site-root view: editor dashboard for editors, public roster otherwise."""

    admin_template = ViewPageTemplateFile('duty_admin.pt')

    def __call__(self):
        portal = _portal(self.context)
        if api.user.has_permission('Modify portal content', obj=portal):
            return self.admin_template()
        return DutyPublicView(self.context, self.request)()


class SelectDutyView(BrowserView):
    """Open an existing roster day or create it, then enter edit mode."""

    def __call__(self):
        request = self.request
        try:
            date_id = _normalise_date(request.form.get('datum'))
            folder = _roster_folder(self.context)
            obj = folder.get(date_id)
            if obj is None:
                obj = api.content.create(
                    container=folder,
                    type=ROSTER_TYPE,
                    id=date_id,
                    title=date_id,
                )
                _publish_if_possible(obj)
            request.response.redirect(obj.absolute_url() + '/edit')
        except Exception as exc:
            api.portal.show_message(str(exc), request=request, type='error')
            request.response.redirect(_portal(self.context).absolute_url() + '/@@dezurstva-admin')
        return ''


class CopyDutyView(BrowserView):
    """Clone duty-roster values from one date to another."""

    def __call__(self):
        request = self.request
        portal = _portal(self.context)
        try:
            source_id = _normalise_date(request.form.get('star'))
            target_id = _normalise_date(request.form.get('nov'))
            folder = _roster_folder(self.context)
            target = folder.get(target_id)
            if target is not None:
                api.portal.show_message(u'Dežurstvo že obstaja.', request=request, type='warning')
                request.response.redirect(target.absolute_url() + '/edit')
                return ''
            source = folder.get(source_id)
            if source is None:
                raise KeyError(u'Izvorno dežurstvo %s ne obstaja.' % source_id)
            values = {name: getattr(source, name)
                      for name in COPY_FIELDS if hasattr(source, name)}
            target = api.content.create(
                container=folder, type=ROSTER_TYPE, id=target_id,
                title=target_id, **values)
            _publish_if_possible(target)
            api.portal.show_message(u'Dežurstvo je bilo uspešno kopirano.', request=request)
            request.response.redirect(folder.absolute_url())
        except Exception as exc:
            api.portal.show_message(str(exc), request=request, type='error')
            request.response.redirect(portal.absolute_url() + '/@@dezurstva-admin')
        return ''
