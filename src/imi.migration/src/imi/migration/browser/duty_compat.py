# -*- coding: utf-8 -*-
"""Compatibility helpers for the Dežurstva public frontend."""
from datetime import date, datetime, timedelta
from io import BytesIO

from plone import api
from Products.Five import BrowserView

from .duty import DutyExportView as BaseDutyExportView
from .duty import DutyHomeView as BaseDutyHomeView
from .duty import DutyPublicView as BaseDutyPublicView
from .duty import ROSTER_TYPE, TEAM_FIELDS, READINESS_FIELDS, _roster_folder


EXPORT_FIELDS = TEAM_FIELDS + (
    ('intervencijska_skupina', u'INTERVENCIJSKA SKUPINA'),
) + READINESS_FIELDS


class DutyPublicView(BaseDutyPublicView):

    def _employee_legacy_id(self, employee):
        return str(getattr(employee, 'source_employee_id', '') or employee.getId())

    def _employee_by_identifier(self, identifier):
        identifier = str(identifier or '')
        directory = self._staff_directory()
        if directory is None:
            return None
        employee = directory.get(identifier)
        if employee is not None:
            return employee
        for candidate in directory.objectValues():
            if getattr(candidate, 'portal_type', None) != 'imi.staff.employee':
                continue
            if self._employee_legacy_id(candidate) == identifier:
                return candidate
        return None

    def staff_title(self, employee_id):
        employee = self._employee_by_identifier(employee_id)
        if employee is not None:
            try:
                return employee.Title() or str(employee_id)
            except Exception:
                pass
        return str(employee_id or '')

    def staff_contact(self, employee_id):
        employee = self._employee_by_identifier(employee_id)
        if employee is None:
            return ''
        return str(getattr(employee, 'contact', '') or '')

    def staff_options(self):
        directory = self._staff_directory()
        if directory is None:
            return []
        result = []
        for employee in directory.objectValues():
            if getattr(employee, 'portal_type', None) != 'imi.staff.employee':
                continue
            if bool(getattr(employee, 'exclude_from_nav', False)):
                continue
            employee_id = self._employee_legacy_id(employee)
            result.append((employee_id, employee.Title() or employee_id))
        return sorted(result, key=lambda item: item[1].lower())

    def readiness_rows(self):
        rows = super(DutyPublicView, self).readiness_rows()
        for row in rows:
            row['contact'] = self.staff_contact(row['employee_id'])
        return rows

    def selected_person_title(self):
        employee = self._employee_by_identifier(self.selected_person())
        if employee is not None:
            return employee.Title() or self.selected_person()
        return self.selected_person()

    def _person_matches_roster(self, obj, person):
        for field_name, _label in EXPORT_FIELDS:
            values = tuple(str(v) for v in (getattr(obj, field_name, ()) or ()))
            if person in values:
                return True
        return False

    def person_dates(self):
        person = self.selected_person()
        if not person:
            return []
        rows = []
        brains = self.portal.portal_catalog(
            portal_type=ROSTER_TYPE,
            path='/'.join(self.roster_folder.getPhysicalPath()),
            sort_on='id', sort_order='ascending')
        for brain in brains:
            obj = brain.getObject()
            if not self._person_matches_roster(obj, person):
                continue
            try:
                parsed = datetime.strptime(obj.getId(), '%Y-%m-%d').date()
            except ValueError:
                continue
            rows.append({
                'id': obj.getId(), 'date': parsed, 'year': parsed.year,
                'month': parsed.month,
                'label': '%02d.%02d.%04d' % (parsed.day, parsed.month, parsed.year),
            })
        return rows

    def today_id(self):
        return date.today().strftime('%Y-%m-%d')

    def easyform_url(self):
        return self.portal.absolute_url() + '/@@dezurstva-change-request'


class DutyHomeView(BaseDutyHomeView):
    def __call__(self):
        portal = api.portal.get()
        if api.user.has_permission('Modify portal content', obj=portal):
            return self.admin_template()
        return DutyPublicView(self.context, self.request)()


class DutyChangeRequestView(BrowserView):
    """Render the EasyForm in the legacy public Dežurstva shell."""

    def __call__(self):
        form = api.portal.get().get('spremeni-dezurstvo')
        if form is None:
            self.request.response.setStatus(404)
            return u'Obrazec Sprememba dežurstva ne obstaja.'

        from collective.easyform.browser.view import EasyFormFormEmbedded

        class PublicEasyForm(EasyFormFormEmbedded):
            def action(inner_self):
                base = api.portal.get().absolute_url() + '/@@dezurstva-change-request'
                datum = str(inner_self.request.form.get('datum') or '').strip()
                return base + (('?datum=' + datum) if datum else '')

        embedded = PublicEasyForm(form, self.request)
        embedded.update()
        body = embedded.render()
        portal_url = api.portal.get().absolute_url()
        css_url = portal_url + '/++resource++imi.migration/duty-public.css'
        easyform_css = portal_url + '/++resource++easyform.css'
        logo_url = portal_url + '/logo.png'
        home_url = portal_url + '/@@dezurstva-public'
        return u'''<!DOCTYPE html>
<html><head><meta charset="utf-8" /><meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Sprememba dežurstva</title>
<link rel="stylesheet" href="%s" />
<link rel="stylesheet" href="%s" />
</head><body class="imi-duty-public imi-duty-change-request">
<div id="okolo"><div id="okol">
<a href="%s"><img id="logo" src="%s" alt="IMI" /></a>
<h1><a href="%s"> DEŽURSTVA IN PRIPRAVLJENOST</a></h1>
<div class="legacy-easyform-wrap">%s</div>
<p class="change-request-back"><a href="%s">Nazaj</a></p>
</div></div></body></html>''' % (
            css_url, easyform_css, home_url, logo_url, home_url, body, home_url)


class DutyExportView(BaseDutyExportView):

    def _parse_date(self, value):
        value = str(value or '').strip()
        if not value:
            raise ValueError('Pri izbiri od - do sta oba datuma obvezna.')
        for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                pass
        raise ValueError('Datum mora biti v obliki DD.MM.LLLL ali LLLL-MM-DD.')

    def _parse_period(self):
        request = self.request.form
        mode = str(request.get('izbira') or request.get('izvoz-1-izbira') or
                   request.get('izvoz-2-izbira') or '')
        today = date.today()
        monday = today - timedelta(days=today.weekday())
        if mode == 'teden':
            return monday - timedelta(days=7), monday - timedelta(days=1)
        if mode == 'tedenplus2':
            return monday, monday + timedelta(days=14)
        if mode == 'tekoci':
            first = today.replace(day=1)
            next_month = (first.replace(day=28) + timedelta(days=4)).replace(day=1)
            return first, next_month - timedelta(days=1)
        if mode == 'pretekli':
            first_this = today.replace(day=1)
            last_previous = first_this - timedelta(days=1)
            return last_previous.replace(day=1), last_previous
        if mode == 'oddo':
            return self._parse_date(request.get('od')), self._parse_date(request.get('do'))
        raise ValueError('Izberite obdobje za izvoz.')

    def _employee_by_identifier(self, directory, identifier):
        identifier = str(identifier or '')
        if directory is None:
            return None
        employee = directory.get(identifier)
        if employee is not None:
            return employee
        for candidate in directory.objectValues():
            if getattr(candidate, 'portal_type', None) != 'imi.staff.employee':
                continue
            legacy = str(getattr(candidate, 'source_employee_id', '') or candidate.getId())
            if legacy == identifier:
                return candidate
        return None

    def __call__(self):
        from openpyxl import Workbook

        try:
            first_day, last_day = self._parse_period()
        except ValueError as exc:
            api.portal.show_message(str(exc), request=self.request, type='error')
            self.request.response.redirect(api.portal.get().absolute_url() + '/@@dezurstva-public')
            return ''

        if first_day > last_day:
            first_day, last_day = last_day, first_day

        person = str(self.request.form.get('oseba') or
                     self.request.form.get('izvoz-2-oseba') or '')
        portal = api.portal.get()
        folder = _roster_folder(self.context)
        directory = portal.get('seznam_zaposlenih')

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'izpis'
        heading = 'Izpis: %s - %s' % (
            first_day.strftime('%d.%m.%Y'), last_day.strftime('%d.%m.%Y'))
        if person:
            employee = self._employee_by_identifier(directory, person)
            heading += ' za osebo %s' % (employee.Title() if employee is not None else person)
        sheet.cell(row=1, column=1, value=heading)

        row = 4
        brains = portal.portal_catalog(
            portal_type=ROSTER_TYPE,
            path='/'.join(folder.getPhysicalPath()),
            sort_on='id', sort_order='ascending')
        for brain in brains:
            try:
                day = datetime.strptime(brain.id, '%Y-%m-%d').date()
            except ValueError:
                continue
            if day < first_day or day > last_day:
                continue
            obj = brain.getObject()
            for field_name, label in EXPORT_FIELDS:
                values = tuple(str(v) for v in (getattr(obj, field_name, ()) or ()))
                if field_name == 'vpis':
                    values = tuple(reversed(values))
                for employee_id in values:
                    if person and employee_id != person:
                        continue
                    employee = self._employee_by_identifier(directory, employee_id)
                    name = employee.Title() if employee is not None else employee_id
                    row += 1
                    sheet.cell(row=row, column=1, value=day.strftime('%d.%m.%Y'))
                    sheet.cell(row=row, column=2, value=label)
                    sheet.cell(row=row, column=3, value=name)
                    if field_name in dict(READINESS_FIELDS):
                        sheet.cell(row=row, column=4,
                                   value=str(getattr(employee, 'contact', '') or '') if employee is not None else '')

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        filename = 'izpis_%s-%s.xlsx' % (
            first_day.strftime('%d_%m_%Y'), last_day.strftime('%d_%m_%Y'))
        response = self.request.response
        response.setHeader('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response.setHeader('Content-Disposition', 'attachment; filename="%s"' % filename)
        return payload.getvalue()
