# -*- coding: utf-8 -*-
"""Admin-only replacements for the legacy Plone 4 data-import helper types."""
import csv
from io import BytesIO, StringIO
import re

from plone import api
from plone.app.textfield.value import RichTextValue
from Products.Five import BrowserView
from Products.Five.browser.pagetemplatefile import ViewPageTemplateFile


PERSON_FIELDS = (
    'predime', 'ime', 'priimek', 'zaime', 'telefon', 'vuporabi', 'maticna',
    'opis', 'enota', 'lokacija', 'dodatna', 'dect', 'fax', 'mobilna',
    'multitone', 'zunanja', 'enaslov', 'zenaslov', 'star',
)


def _upload_bytes(request, name):
    upload = request.form.get(name)
    if upload is None:
        return b'', ''
    filename = getattr(upload, 'filename', '') or getattr(upload, 'name', '') or ''
    reader = getattr(upload, 'read', None)
    if callable(reader):
        data = reader()
    elif getattr(upload, 'file', None) is not None:
        data = upload.file.read()
    else:
        data = bytes(upload)
    if isinstance(data, str):
        data = data.encode('utf-8')
    return data or b'', str(filename)


def _decode_csv(data):
    for encoding in ('utf-8-sig', 'cp1250', 'latin-1'):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode('utf-8', 'replace')


def _safe_id(value):
    # Preserve the old importer contract: only ASCII letters/digits/hyphens.
    return ''.join(ch for ch in str(value or '') if ch.isascii() and (ch.isalnum() or ch == '-'))


def _person_id(name, surname, source_id):
    return _safe_id('%s_%s_%s' % (name, surname, source_id)).replace('_', '')


def _publish_if_possible(obj):
    try:
        if api.content.get_state(obj=obj) == 'published':
            return
        transitions = api.content.get_transitions(obj=obj)
        if 'publish' in [item.get('id') for item in transitions]:
            api.content.transition(obj=obj, transition='publish')
    except Exception:
        pass


def _find_by_id(site, portal_type, obj_id):
    brains = site.portal_catalog(
        portal_type=portal_type,
        id=obj_id,
        path='/'.join(site.getPhysicalPath()),
    )
    return brains[0].getObject() if brains else None


def _rich(value):
    return RichTextValue(
        raw=str(value or ''), mimeType='text/html',
        outputMimeType='text/x-html-safe')


def _cell(value):
    if value is None:
        return ''
    if isinstance(value, float):
        # xlrd returned numeric cells as floats; retain that legacy spelling.
        return str(value)
    return str(value)


def _workbook_rows(data, filename):
    lower = filename.lower()
    if lower.endswith('.xlsx') or lower.endswith('.xlsm'):
        from openpyxl import load_workbook
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        return [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
    if lower.endswith('.xls'):
        import xlrd
        workbook = xlrd.open_workbook(file_contents=data)
        sheet = workbook.sheet_by_index(0)
        return [[sheet.cell_value(r, c) for c in range(sheet.ncols)]
                for r in range(sheet.nrows)]
    raise ValueError('Podprte so datoteke .xls, .xlsx in .xlsm.')


class DirectoryImportView(BrowserView):
    template = ViewPageTemplateFile('directory_import.pt')

    def __init__(self, context, request):
        super().__init__(context, request)
        self.report = None
        self.error = None

    def __call__(self):
        if self.request.method == 'POST':
            try:
                self.report = self._import()
                api.portal.show_message(
                    'Uvoz imenika je končan: %d novih, %d posodobljenih, %d preskočenih, %d napak.' % (
                        self.report['created'], self.report['updated'],
                        self.report['skipped'], len(self.report['errors'])),
                    request=self.request)
            except Exception as exc:
                self.error = str(exc)
                api.portal.show_message('Uvoz imenika ni uspel: %s' % exc,
                                        request=self.request, type='error')
        return self.template()

    def _folder(self, base, folder_id, title):
        folder_id = _safe_id(folder_id)
        if not folder_id:
            raise ValueError('Prazna oznaka oddelka/mapa.')
        if folder_id in base.objectIds():
            return base[folder_id]
        obj = api.content.create(container=base, type='Folder', id=folder_id,
                                 title=title or folder_id, safe_id=False)
        _publish_if_possible(obj)
        return obj

    def _import(self):
        people_data, people_name = _upload_bytes(self.request, 'datoteka')
        structure_data, structure_name = _upload_bytes(self.request, 'datoteka2')
        if not people_data:
            raise ValueError('Izberite Datoteko s podatki imenika.')
        site = api.portal.get()
        base = site.get('data2')
        if base is None:
            raise ValueError('Manjka /portal/data2.')

        parent_map = {}
        if structure_data:
            rows = list(csv.reader(StringIO(_decode_csv(structure_data)), delimiter=';', quotechar='"'))
            for row in rows[1:]:
                if len(row) < 3:
                    continue
                code, title = row[0].strip(), row[2].strip()
                folder_id = _safe_id(title.replace(' ', '-'))
                if code and folder_id:
                    parent_map[code] = (folder_id, title)
                    self._folder(base, folder_id, title)

        rows = list(csv.reader(StringIO(_decode_csv(people_data)), delimiter=';', quotechar='"'))
        created = updated = skipped = 0
        errors = []
        current_folder = base
        update_existing = str(self.request.form.get('update_existing', '')).lower() in ('1', 'true', 'on', 'yes')

        for index, row in enumerate(rows):
            try:
                if not row:
                    continue
                first = row[0].strip()
                if first == '#ODDELEK':
                    title = row[1].strip() if len(row) > 1 else ''
                    lookahead_code = ''
                    if index + 3 < len(rows) and rows[index + 3]:
                        lookahead_code = rows[index + 3][0].strip()
                    folder_id = _safe_id(title.replace(' ', '-'))
                    if lookahead_code in parent_map:
                        parent_id, parent_title = parent_map[lookahead_code]
                        parent = self._folder(base, parent_id, parent_title)
                        current_folder = self._folder(parent, folder_id, title)
                    else:
                        current_folder = self._folder(base, folder_id, title)
                    continue
                if first.startswith('#'):
                    continue
                if len(row) < 22:
                    continue

                source_id = row[1].strip()
                values = {
                    'vuporabi': row[3].strip(), 'maticna': row[4].strip(),
                    'predime': row[5].strip(), 'ime': row[6].strip(),
                    'priimek': row[7].strip(), 'zaime': row[8].strip(),
                    'opis': row[9].strip(), 'enota': row[10].strip(),
                    'lokacija': row[11].strip(), 'telefon': row[12].strip(),
                    'dodatna': row[13].strip(), 'dect': row[14].strip(),
                    'fax': row[15].strip(), 'multitone': row[16].strip(),
                    'mobilna': row[17].strip(), 'zunanja': row[18].strip(),
                    'enaslov': row[19].strip(), 'zenaslov': row[20].strip(),
                    'star': row[21].strip(),
                }
                obj_id = _person_id(values['ime'], values['priimek'], source_id)
                if not obj_id:
                    raise ValueError('Ni mogoče sestaviti ID-ja osebe.')
                existing = _find_by_id(site, 'imi.directory.person', obj_id)
                if existing is not None and not update_existing:
                    skipped += 1
                    continue
                if existing is None:
                    existing = api.content.create(
                        container=current_folder, type='imi.directory.person',
                        id=obj_id,
                        title='%s %s %s' % (values['ime'], values['priimek'], source_id),
                        safe_id=False)
                    created += 1
                else:
                    updated += 1
                for field in PERSON_FIELDS:
                    setattr(existing, field, values[field])
                existing.oddelek = current_folder.Title() if current_folder is not base else ''
                _publish_if_possible(existing)
                existing.reindexObject()
            except Exception as exc:
                errors.append('vrstica %d: %s' % (index + 1, exc))

        import transaction
        transaction.commit()
        return {
            'filename': people_name,
            'structure_filename': structure_name,
            'created': created, 'updated': updated, 'skipped': skipped,
            'errors': errors,
        }


class ExamsImportView(BrowserView):
    template = ViewPageTemplateFile('exams_import.pt')

    def __init__(self, context, request):
        super().__init__(context, request)
        self.report = None
        self.error = None

    def __call__(self):
        if self.request.method == 'POST':
            try:
                self.report = self._import()
                api.portal.show_message(
                    'Uvoz preiskav je končan: %d novih, %d posodobljenih, %d napak.' % (
                        self.report['created'], self.report['updated'], len(self.report['errors'])),
                    request=self.request)
            except Exception as exc:
                self.error = str(exc)
                api.portal.show_message('Uvoz preiskav ni uspel: %s' % exc,
                                        request=self.request, type='error')
        return self.template()

    def _container(self, site):
        current = site
        for obj_id in ('preiskave-1', 'katalog-preiskav'):
            current = current.get(obj_id) if current is not None else None
        if current is None:
            raise ValueError('Manjka /preiskave/preiskave-1/katalog-preiskav.')
        return current

    def _import(self):
        exams_data, exams_name = _upload_bytes(self.request, 'datoteka1')
        samples_data, samples_name = _upload_bytes(self.request, 'datoteka2')
        if not exams_data or not samples_data:
            raise ValueError('Izberite obe datoteki: preiskave in vzorci.')
        exam_rows = _workbook_rows(exams_data, exams_name)
        sample_rows = _workbook_rows(samples_data, samples_name)
        if len(exam_rows) < 2 or len(sample_rows) < 2:
            raise ValueError('Excel datoteki nimata podatkovnih vrstic.')

        samples = {}
        exams = {}
        for raw in exam_rows[1:]:
            row = list(raw) + [''] * max(0, 31 - len(raw))
            row = [_cell(value) for value in row]
            code = row[0]
            if not code:
                continue
            sample_code = row[15]
            if code not in exams:
                exams[code] = {
                    'podrocje': [], 'sifra': code, 'title': row[1], 'metode': row[6],
                    'opis': row[7], 'sinonim': row[4], 'trajanje': row[9],
                    'urnik': row[10], 'opombe': row[8], 'sklop': row[3],
                    'sample_codes': [], 'vzorci_lab_kratica': [], 'vzorci_lab': [],
                    'vzorci_lab_tel': [], 'vzorci_odvzem': [],
                    'vzorci_odvzem_video_sifra': [], 'vzorci_kolicina': [],
                    'embalaza_slika_sifra': [], 'vzorci_transport': [],
                    'vzorci_opomba': [], 'laboratoriji': row[5], 'link_sop': row[11],
                    'nova_pre': row[12], 'nujna_pre': row[13],
                    'preiskava_vzorec_nujno': [],
                }
            item = exams[code]
            item['podrocje'].append(row[2])
            item['sample_codes'].append(sample_code)
            item['vzorci_lab_kratica'].append('%s###%s###' % (row[17], row[20]))
            item['vzorci_lab'].append('%s###%s###' % (row[18], row[21]))
            item['vzorci_lab_tel'].append('%s###%s###%s' % (row[19], row[22], row[24]))
            item['vzorci_odvzem'].append(row[25])
            item['vzorci_odvzem_video_sifra'].append(row[26])
            item['vzorci_kolicina'].append(row[27])
            item['embalaza_slika_sifra'].append(row[28])
            item['vzorci_transport'].append(row[29])
            item['vzorci_opomba'].append(row[30])
            item['preiskava_vzorec_nujno'].append(row[16])
            if sample_code:
                samples[sample_code] = row[14]

        for raw in sample_rows[1:]:
            row = list(raw) + [''] * max(0, 7 - len(raw))
            row = [_cell(value) for value in row]
            code = row[0]
            if not code:
                continue
            base = samples.get(code, '')
            samples[code] = base + '######' + '###'.join(row[1:7])

        site = api.portal.get()
        container = self._container(site)
        created = updated = 0
        errors = []
        for code, item in exams.items():
            try:
                obj_id = 'preiskava_' + code
                obj = _find_by_id(site, 'imi.exams.examination', obj_id)
                if obj is None:
                    obj = api.content.create(container=container, type='imi.exams.examination',
                                             id=obj_id, title=item['title'] or obj_id,
                                             safe_id=False)
                    created += 1
                else:
                    updated += 1
                sample_values = []
                levels = [[] for _ in range(6)]
                for sample_code in item['sample_codes']:
                    value = samples.get(sample_code, '')
                    sample_values.append(value)
                    parts = value.split('###')
                    for index in range(6):
                        levels[index].append(parts[index + 2] if len(parts) > index + 2 else '')

                obj.title = item['title'] or obj_id
                obj.sifra = code
                obj.podrocje = tuple(item['podrocje'])
                obj.metode = _rich(item['metode'])
                obj.opis = _rich(item['opis'])
                obj.sinonim = _rich(item['sinonim'])
                obj.trajanje = _rich(item['trajanje'])
                obj.urnik = _rich(item['urnik'])
                obj.opombe = _rich(item['opombe'])
                obj.sklop = _rich(item['sklop'])
                obj.laboratoriji = item['laboratoriji']
                obj.link_sop = item['link_sop']
                obj.nova_pre = item['nova_pre']
                obj.nujna_pre = item['nujna_pre']
                obj.vzorci = tuple(sample_values)
                obj.preiskava_vzorec_nujno = tuple(item['preiskava_vzorec_nujno'])
                for field in ('vzorci_lab_kratica', 'vzorci_lab', 'vzorci_lab_tel',
                              'vzorci_odvzem', 'vzorci_odvzem_video_sifra',
                              'vzorci_kolicina', 'embalaza_slika_sifra',
                              'vzorci_transport', 'vzorci_opomba'):
                    setattr(obj, field, tuple(item[field]))
                for index in range(6):
                    setattr(obj, 'vzorci_nivo%d' % (index + 1), tuple(levels[index]))
                _publish_if_possible(obj)
                obj.reindexObject()
            except Exception as exc:
                errors.append('%s: %s' % (code, exc))

        import transaction
        transaction.commit()
        return {
            'exams_filename': exams_name, 'samples_filename': samples_name,
            'exam_rows': len(exam_rows) - 1, 'sample_rows': len(sample_rows) - 1,
            'created': created, 'updated': updated, 'errors': errors,
        }
